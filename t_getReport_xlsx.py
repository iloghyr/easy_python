#c70e5e7fdf776b48de33312777c9e731
#coding=utf-8
import urllib2
import urllib
import json
import xlrd
import sys
from openpyxl.workbook import Workbook  
from openpyxl.writer.excel import ExcelWriter 
import time
import cookielib

start = time.time()
#设置系统编码
reload(sys)
sys.setdefaultencoding( "utf-8" )
filename = r'jinshan.xlsx'
wb = Workbook()
ew = ExcelWriter(workbook = wb)
ws = wb.worksheets[0]
ws.cell(row=0,column=0).value = '基本信息'
ws.cell(row=0,column=1).value = '危险行为'
ws.cell(row=0,column=2).value = '其他行为'
ws.cell(row=0,column=3).value = '权限列表'
ws.cell(row=0,column=4).value = '启动方式'
ws.cell(row=0,column=5).value = '广告相关'
ws.cell(row=0,column=6).value = '文件操作'
ws.cell(row=0,column=7).value = '网络监控'
cur_row =  ws.get_highest_row()
now_total = 0
print 'current row:',cur_row

def doWrite(row,col,cont):
	global ws
	ws.cell(row=row,column=col).value=cont
#模拟登入
def doLogin(user):
	head = {
    'Referer':'https://login.ijinshan.com/fireeye/login.html?service=https%3A%2F%2Ffireeye.ijinshan.com%2Faccount%2Flogin',
    'Host':'login.ijinshan.com',
    'Connection':'Connection',
    'Accept-Language':'zh-cn,zh;q=0.8,en-us;q=0.5,en;q=0.3',
    'Accept-Encoding':'gzip, deflate',
    'Accept':'*/*',
    'User-Agent':'Mozilla/5.0 (Windows NT 6.1; rv:19.0) Gecko/20100101 Firefox/19.0',
    'X-Requested-With':'XMLHttpRequest'
	}
	str1 = str(time.time())
	url = 'https://login.ijinshan.com/glt?_lt='+str1
	r = urllib2.Request(url,headers=head)
	ckjar =  cookielib.CookieJar()
	opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(ckjar))
	res = opener.open(r)
	#print res.info()
	print '------------'
	#登入 step 1
	loginurl = 'https://login.ijinshan.com/login'
	data = {
	    'cc':'',	
	    'cn':'e6c5aeba967a9e286b083044629c6091',
	    'pwd':'bf8fcf8f58466941d03091f73ff46e0c',
	    'rm':'1',
	    'service':'https://fireeye.ijinshan.com/account/login',
	    'user':''
	}
	data['user'] = user
	data = urllib.urlencode(data)
	res = opener.open(loginurl,data)
	res_json =  json.loads(res.read().strip())
	#登入 step 2 并获得最后登入成功后的cookie
	url = res_json['url']
	#获取session
	res = opener.open(url)
	return opener

#=========================end def function ==========================

users  = []
for line in open('user'):
	users.append(line.strip())

for u in range(len(users)):
	doOpener = doLogin(users[u])
	print '[Dealing]: ',u
	md5_list = []
	sha1_list = []
	
	#==========================获取列表==========================
	url = 'http://fireeye.ijinshan.com/secure/VsampleBehavStat?pageSize=10000'
	print 'get list: ',
	res = doOpener.open(url).read()
	res_json = json.loads(res)
	try:
		for i in res_json['store']:
			md5_list.append(i['md5'])
			sha1_list.append(i['sha1'])
	except:
		continue
	print ' done',
	#head['Cookie'] = '__utma=214684160.447770654.1365401413.1365564436.1365575572.9; __utmz=214684160.1365401413.1.1.utmcsr=(direct)|utmccn=(direct)|utmcmd=(none); __utmc=214684160; JSESSIONID=aaaRuM7aQ4JbGbN5xYX3t; __utmb=214684160.6.10.1365575572'
	for i in range(len(md5_list)):
		if now_total == 100:  #control num
			pass
		while True:
			try:
				now_total += 1
				print 'Dealing ',(i+1),' of ',len(md5_list),' Total:',now_total,'  user:',users[u]
				md5 = md5_list[i]
				sha1 = sha1_list[i]
				
				#print 'dealing :',md5
				
			
				#==========================基本信息==========================
				url = 'http://fireeye.ijinshan.com/secure/isAnalysised?md5='+md5+'&sha1='+sha1
			
				print 'basic info: ',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print res_json  
				#print json.dumps(res_json,ensure_ascii=False)
				App_name = ''
				App_version = ''
				try: 
					App_name =  res_json['store'][0]['appName']
				except:
					print 'appName Not Define'
				try:
					App_md5 =  res_json['store'][0]['md5']
					App_sha1 = res_json['store'][0]['sha1']
				except:
					#continue
					break
				try: 
					App_version = res_json['store'][0]['version']
				except:
					print 'version Not Define'
				cont0 = '文件名：'+App_name+'\nmd5='+App_md5+'\nsha1='+App_sha1+'\n版本：'+str(App_version)
				
				print ' done',
			
				#==========================权限列表==========================
				url = 'http://fireeye.ijinshan.com/secure/apkReport?md5='+md5+'&sha1='+sha1+'&action=apkpermop'
				print 'quanxian:',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print json.dumps(res_json,ensure_ascii=False)
				cont3 = ''
				if res_json['header'] == '800':
					for res in res_json['store']:
						isused = '已使用' if res['is_used'] == 1 else '未使用'
						cont3 += res['permDesc']+'  '+res['permName']+'  '+isused+'\n'
					#print cont
				
				print 'done ',
			
			
			
				#==========================广告相关==========================
				url = 'http://fireeye.ijinshan.com/secure/apkReport?md5='+md5+'&sha1='+sha1+'&action=apkadop'
				print 'ad:',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print json.dumps(res_json,ensure_ascii=False)
				cont5 = ''
				if res_json['header'] == '800':
					for res in res_json['store']:
						cont5 += res['adName']+'('+res['adDesc']+')\n'
					#print cont
				
				print 'done ',
			
				#==========================危险行为==========================
				url = 'http://fireeye.ijinshan.com/secure/apkReport?md5='+md5+'&sha1='+sha1+'&action=apkbehva&blevel=1'
				print 'dangerbehva :',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print json.dumps(res_json,ensure_ascii=False)
				cont1 = ''
				if res_json['header'] == '800':
						for res in res_json['store']:
							cont1 += '行为描述:'+res['behavDesc']+'\n附加信息:'+res['content']+'\n\n'
				
				print 'done ',
			
			
				#==========================其他行为==========================
				url = 'http://fireeye.ijinshan.com/secure/apkReport?md5='+md5+'&sha1='+sha1+'&action=apkbehva&blevel=2'
				print 'other  :',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print json.dumps(res_json,ensure_ascii=False)
				cont2 = ''
				if res_json['header'] == '800':
					for res in res_json['store']:
						detail = ''
						try:
							for d1 in res['contents']:
								detail += d1+'\n'
							cont2 += '行为描述:'+res['behavDesc']+'\n详细:\n'+detail+'\n\n'
						except:
							pass	
					#print cont
				
				print 'done ',
			
			
			
				#==========================文件操作==========================
				url = 'http://fireeye.ijinshan.com/secure/apkReport?md5='+md5+'&sha1='+sha1+'&action=apkfileop'
				print 'file :',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print json.dumps(res_json,ensure_ascii=False)
				cont6 = ''
				if res_json['header'] == '800':
					for i in res_json['store'][0]:
						for j in res_json['store'][0][i]:
							cont6 += i+'  '+j['tgtFile']+'\n'
				#print cont 
			
				
				print 'done ',
			
			
			
				#==========================启动方式==========================
				url = 'http://fireeye.ijinshan.com/secure/apkReport?md5='+md5+'&sha1='+sha1+'&action=apkstarttype'
				print 'apk start :',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print json.dumps(res_json,ensure_ascii=False)
				cont4 = ''
				if res_json['header'] == '800':
					for i in res_json['store']:
						cont4 += '启动方式：'+i['starttypeDesc']+'\n启动对象：'+i['content']+'\n\n'
			
				#print cont 
			
				
				print 'done ',
			
			
				#==========================网络监控==========================
				url = 'http://fireeye.ijinshan.com/secure/apkReport?md5='+md5+'&sha1='+sha1+'&action=apknetop'
				print 'net start :',
				res = doOpener.open(url).read()
				res_json = json.loads(res)
				#print json.dumps(res_json,ensure_ascii=False)
				cont7 = ''
				if res_json['header'] == '800':
					for i in res_json['store']:
						method = '[POST]:' if i['opType']=='0' else '[GET]:'
						cont7 += method+str(i['tgtUrl'])+'\n'
				#print cont
				print 'done'
				isAllDone = True 
			except Exception,e:
				print e
				isAllDone = False
				print '[ERROR]...............................................Retry'
			print '**********************************************************************'
			if isAllDone:
				doWrite(cur_row,0,cont0)
				doWrite(cur_row,1,cont1)
				doWrite(cur_row,2,cont2)
				doWrite(cur_row,3,cont3)
				doWrite(cur_row,4,cont4)
				doWrite(cur_row,5,cont5)
				doWrite(cur_row,6,cont6)
				doWrite(cur_row,7,cont7)
				cur_row += 1
				doOpener.close()
				break





ew.save(filename)
end = time.time()
print 'Total time:',end-start,'s'
print 'SUCCESS'